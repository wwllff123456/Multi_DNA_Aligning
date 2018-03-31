# included packages in python itself:
import tkinter as tk
from tkinter import messagebox, simpledialog, filedialog
import os
import datetime
import statistics
import subprocess
import winsound
import logging
import xml.etree.ElementTree as ET

# packages that may need to be installed:

try:
    import numpy as np
except ImportError:
    a = tk.Tk()
    a.withdraw()
    winsound.PlaySound("SystemHand", winsound.SND_ALIAS)
    ans = messagebox.askyesno('Error', 'Error occureed during importing numpy,\nplease install/re-install this module.\n\nCopy command to clipboard?')
    print(ans)
    if ans:
        os.system("echo '%s' | clip" % 'pip install numpy')
    os._exit(1)

try:
    import openpyxl as xl
except ImportError:
    a = tk.Tk()
    a.withdraw()
    winsound.PlaySound("SystemHand", winsound.SND_ALIAS)
    ans = messagebox.askyesno('Error','Error occureed during importing openpyxl,\nplease install/re-install this module.\n\nCopy command to clipboard?')
    if ans:
        os.system("echo '%s' | clip" % 'pip install openpyxl')
    os._exit(1)


# these are for debug purpose:

# import matplotlib.pyplot as plt
# import math
# import time


# global variable settings:
    # for ranking, will NOT work if larger than smallest data count in one plate
get_largest = 8

    # for the std starting position:
std_position = 0

    # for bg position
bg_position = 10

class Well:

    def __init__(self):

        self.well_id = ''
        self.name = ''
        self.row = ''
        self.col = ''

        self.in_plate = ''

        self.raw_data = None

        self.y_original = None

        self.y_bg = None              # y value after bg removed

        self.x_fit = None             # fitted x using y value

        self.x_calc = None            # x calc = 0.1 - x

        self.z_score = None           # z score = (x_calc - average_of_x_calc) / dev_of_x_calc

    def y(self):
        try:
            self.y_original = float(self.raw_data)
        except ValueError:
            self.y_original = '#ERR'

    def remove_bg(self, bg):
        try:
            self.y_bg = self.y_original - bg
        except TypeError:
            self.y_bg = '#ERR'

    def fit_x(self, a, b):

        # given y = ax + b
        # x = (y - b) / a

        try:
            self.x_fit = (self.y_bg - b) / a
        except TypeError:
            self.x_fit = '#ERR' \

    def calc_x(self):

        # x calc = 0.1 - x

        self.x_calc = 0.1 - self.x_fit

    def calc_z(self, mean, dev):

        try:
            self.z_score = (self.x_calc - mean) / dev
        except TypeError:
            self.z_score = '#ERR'



class Plate:

    bg = None                       # bg that needs to be removed for y value
    bg_available = False

    a = None                        # y = a * x + b
    b = None

    def __init__(self):

        self.name = ''
        self.read_time = ''

        self.wells = []

        self.bg_removed = False     # bg removed or not for y value
        self.x_calc_mean = None
        self.x_calc_dev = None

        self.group_num = 0

    def load_wells(self, well_list):
        self.wells = well_list

    def set_group(self, n):
        self.group_num = n

    def set_name(self, name):
        self.name = name

    def remove_bg(self):
        if self.wells and not self.bg_removed:
            for well in self.wells:
                well.remove_bg(self.bg)
            self.bg_removed = True

    def x_fit(self):
        if self.wells:
            for well in self.wells:
                well.fit_x(self.a, self.b)

    def calc_x(self):
        if self.wells:
            for well in self.wells:
                well.calc_x()

    def calc_x_mean(self):
        # print([i.x_calc for i in self.wells])
        self.x_calc_mean = statistics.mean(filter(lambda x: x != '#ERR', [i.x_calc for i in self.wells]))

    def calc_x_dev(self):
        self.x_calc_dev = statistics.stdev(filter(lambda x: x != '#ERR', [i.x_calc for i in self.wells]))

    def calc_z(self):
        self.calc_x_mean()
        self.calc_x_dev()

        if self.wells:
            for well in self.wells:
                well.calc_z(self.x_calc_mean, self.x_calc_dev)

    def quick_run(self):
        self.remove_bg()
        self.x_fit()
        self.calc_x()
        self.calc_z()


class STD(Plate):

    @classmethod
    def set_bg(cls, bgg):
        Plate.bg = bgg
        Plate.bg_available = True

    @classmethod
    def set_fit(cls, aa, bb):
        Plate.a = aa
        Plate.b = bb

    def __init__(self):
        Plate.__init__(self)

        self.std_x = []
        self.std_y = []                  # this y is after bg removing!

    def set_std_x(self, xlist):
        self.std_x = xlist

    def load_bg(self):
        global bg_position
        self.set_bg(self.wells[bg_position].y_original)

    def load_std_y(self):               # this needs to be ran after setting std_x
        global std_position
        self.std_y = [self.wells[i].y_bg for i in range(std_position, std_position+len(self.std_x))]

    def fit(self):
        x = np.array(self.std_x)
        y = np.array(self.std_y)

        print('fitting with x:\n{}\ny:\n{}'.format(x, y))

        aa, bb = np.polyfit(x, y, 1)

        self.set_fit(aa, bb)

    def std_quick_run(self):
        self.load_bg()
        self.remove_bg()
        self.load_std_y()
        self.fit()


class AllPlates:

    def __init__(self):

        self.std = STD()
        self.plates = []

        self.project_name = ''

        self.std_path = ''
        self.data_path = ''

        self.groups = []
        self.group_settings = []
        self.ranked_groups = []

    def load_std(self):

        with open(self.std_path, 'r') as std:
            for _ in range(20):
                a = std.readline()
                if a:
                    break
            else:
                messagebox.showerror('Error', 'Not a valid std file\nEmpty file?')
                return
            # print(repr(a))
            if a:
                values = a.lstrip(',').rstrip(',').rstrip('\n').replace(' ', '').split(',')
                for i in range(len(values)):
                    try:
                        values[i] = float(values[i])
                    except ValueError:
                        messagebox.showerror('Error', 'Not a valid std file')
                        return
                else:
                    std_x = values[:]

        self.std.set_std_x(std_x)       # writes the std file x values into std obj

    def load_data(self):                # this must run after load_std()

        p = ET.parse(self.data_path)
        self.project_name = os.path.splitext(os.path.basename(self.data_path))[0]
        root = p.getroot()

        for platesections in root:
            for platesection in platesections:
                platesection_dict = platesection.attrib
                plate_name = platesection_dict['Name']
                read_time = platesection_dict['ReadTime']
                print(plate_name, read_time, end='\t\t\t')

                wavelengths = platesection.find('Wavelengths')
                for wavelength in wavelengths:
                    for wells in wavelength:
                        well_list = []
                        for well in wells:
                            well_dict = well.attrib
                            data = well.find('RawData').text

                            w = Well()
                            w.in_plate = plate_name
                            w.well_id = well_dict['WellID']
                            w.name = well_dict['Name']
                            w.row = well_dict['Row']
                            w.col = well_dict['Col']
                            w.raw_data = data
                            w.y()

                            well_list.append(w)

            if plate_name.lower()[:3] == 'std':
                print(plate_name, 'is an STD plate!!!')
                p = STD()
                self.std = p
            else:
                print(plate_name, 'is a normal plate!!!')
                p = Plate()

            p.name = plate_name
            p.wells = well_list
            self.plates.append(p)

    def set_group(self, glist):
        assert len(glist) == len(self.plates)

        self.group_settings = glist

        for i in range(len(glist)):
            self.plates[i].group_num = glist[i]

    def calc(self):
        self.std.std_quick_run()
        for i in self.plates:
            i.quick_run()

    def rank(self):
        self.group_settings = [i.group_num for i in self.plates]
        grouped = max(self.group_settings)

        all_plates = self.plates[:]
        print(len(all_plates))

        if grouped > 0:
            for i in range(1, grouped+1):
                for plate in all_plates:
                    if plate.group_num == 0:
                        self.groups.append([plate])
                    else:
                        grouped_plates = [i for i in filter(lambda x: x.group_num == i, self.plates)]
                        self.groups.append(grouped_plates)
                        break
                for i in self.groups:
                    for j in i:
                        try:
                            all_plates.remove(j)
                        except ValueError:
                            continue
        if all_plates:
            for i in all_plates:
                print(i.name)
                self.groups.append([i])

        for group in self.groups:
            names = [i.name for i in group]
            wells = []
            for plate in group:
                wells += plate.wells
            ranked = sorted(wells, key=lambda x: x.z_score)[::-1]
            self.ranked_groups.append([names, ranked])

        global get_largest
        for i in self.ranked_groups:
            print(i[0])
            for j in range(get_largest):
                print(i[1][j].z_score, sep='\t')

    def export(self):

        self.excel_path = os.path.splitext(self.data_path)[0] + '_export' + '.xlsx'

        export_xlsx = xl.Workbook()
        ws = export_xlsx.active
        ws.title = 'export'
        time = '{:%Y-%m-%d %H:%M:%S}'.format(datetime.datetime.now())
        ws.append([self.project_name, '', 'exported at', '', time])
        ws.append([])
        ws.append([])
        ws.append(['bg', self.std.bg])
        ws.append([])
        ws.append(['std x'] + self.std.std_x)
        ws.append([])
        ws.append(['std y'] + self.std.std_y)
        ws.append([])
        ws.append(['a:', self.std.a, '', '', 'b:', self.std.b])

        for plate in self.plates:
            new_sheet = export_xlsx.create_sheet()
            new_sheet.title = plate.name

            new_sheet.append(['ID', 'Name', 'Row', 'Col', 'y', 'y-bg', 'x(fit)', 'x(calc)', 'z score', '', '',
                              'x mean:', plate.x_calc_mean, '', 'x dev', plate.x_calc_dev])

            for well in plate.wells:
                new_sheet.append([well.well_id, well.name, well.row, well.col, well.y_original,
                                  well.y_bg, well.x_fit, well.x_calc, well.z_score])

        ranked = export_xlsx.create_sheet()
        ranked.title = 'Ranked'

        global get_largest

        for i in range(len(self.ranked_groups)):
            ranked.cell(column=5 * i + 1, row=1, value='z score')
            ranked.cell(column=5 * i + 2, row=1, value='x calc')
            ranked.cell(column=5 * i + 3, row=1, value='in plate')
            ranked.cell(column=5 * i + 4, row=1, value='name')

            for k in range(get_largest):
                ranked.cell(column=5 * i + 1, row=len(self.ranked_groups[i][0]) + 2 + k,
                            value=self.ranked_groups[i][1][k].z_score)
                ranked.cell(column=5 * i + 2, row=len(self.ranked_groups[i][0]) + 2 + k,
                            value=self.ranked_groups[i][1][k].x_calc)
                ranked.cell(column=5 * i + 3, row=len(self.ranked_groups[i][0]) + 2 + k,
                            value=self.ranked_groups[i][1][k].in_plate)
                ranked.cell(column=5 * i + 4, row=len(self.ranked_groups[i][0]) + 2 + k,
                            value=self.ranked_groups[i][1][k].name)

            for j in range(len(self.ranked_groups[i][0])):
                ranked.cell(column=5 * i + 1, row=get_largest + 2, value=self.ranked_groups[i][0][j])

        export_xlsx.save(self.excel_path)


class GUI:

    def __init__(self):

        self.root = tk.Tk()
        # self.root.resizable(0, 0)
        # self.root.title = ('UV Calc')

        self.std_path = 'n/a'
        self.data_path = 'n/a'

        self.std_string = tk.StringVar()
        self.std_string.set(self.std_path)

        self.data_string = tk.StringVar()
        self.data_string.set(self.data_path)

        self.name_string = tk.StringVar()
        self.name_string.set('n/a')

        self.group_result = []

        self.mk_window()
        self.root.mainloop()

    def mk_window(self):

        main_frame = tk.Frame(self.root, width=800, height=800, padx=10, pady=10)
        main_frame.pack(side=tk.TOP, fill=tk.BOTH)

        spacer_top = tk.Frame(main_frame, width=600, height=1)
        spacer_top.pack(side=tk.TOP, fill=tk.X)

        # spacer_left = tk.Frame(main_frame, width=10, height=800)
        # spacer_left.pack(side=tk.LEFT, fill=tk.Y)
        #
        # spacer_right = tk.Frame(main_frame, width=10, height=800)
        # spacer_right.pack(side=tk.RIGHT, fill=tk.Y)
        #
        # spacer_bottom = tk.Frame(main_frame, width=10, height=800)
        # spacer_bottom.pack(side=tk.BOTTOM, fill=tk.X)

        # top frame: the loading area for data and std files.

        top_frame = tk.Frame(main_frame, width=600, height=200, padx=10, pady=10)
        top_frame.pack(side=tk.TOP, fill=tk.X)

        name_label = tk.Label(top_frame, textvariable=self.name_string, font=['arial', '15'])
        name_label.pack(side=tk.TOP, fill=tk.X)

        spacer_top_1 = tk.Frame(top_frame, width=600, height=15)
        spacer_top_1.pack(side=tk.TOP, fill=tk.X)


        # data file:

        data_frame = tk.Frame(top_frame, width=400, height=60, bd=3, relief=tk.GROOVE)
        data_frame.pack(side=tk.TOP, fill=tk.BOTH)

        browse_data_button = tk.Button(data_frame, text='data.txt', command=self.get_data_path)
        browse_data_button.pack(side=tk.LEFT, padx=10, pady=10)

        self.edit_data_button = tk.Button(data_frame, text='edit', state=tk.DISABLED, command=self._edit_data)
        self.edit_data_button.pack(side=tk.RIGHT, padx=10, pady=10)

        self.dir_button = tk.Button(data_frame, text='folder', state=tk.DISABLED, command=self._open_dir)
        self.dir_button.pack(side=tk.RIGHT, padx=10, pady=10)

        data_label = tk.Label(data_frame, textvariable=self.data_string)
        data_label.pack(side=tk.LEFT)

        # std file:

        std_frame = tk.Frame(top_frame, width=400, height=60, bd=3, relief=tk.GROOVE)
        std_frame.pack(side=tk.TOP, fill=tk.BOTH)

        self.browse_std_button = tk.Button(std_frame, text=' std.csv', state=tk.DISABLED, command=self.get_std_path)
        self.browse_std_button.pack(side=tk.LEFT, padx=10, pady=10)

        self.edit_std_button = tk.Button(std_frame, text='edit', state=tk.DISABLED, command=self._edit_std)
        self.edit_std_button.pack(side=tk.RIGHT, padx=10, pady=10)

        self.new_std_button = tk.Button(std_frame, text='new', state=tk.DISABLED, command=self._new_std)
        self.new_std_button.pack(side=tk.RIGHT, padx=10, pady=10)

        std_label = tk.Label(std_frame, textvariable=self.std_string)
        std_label.pack(side=tk.LEFT)

        # middle frame: all functions in here

        middle_frame = tk.Frame(main_frame, width=600, height=200, padx=10, pady=10)
        middle_frame.pack(side=tk.TOP, anchor=tk.CENTER)

        self.load_button = tk.Button(middle_frame, text='Load\ndata', width=10, height=2, state=tk.DISABLED, command=self.load)
        self.load_button.grid(row=2, column=1, padx=20, pady=10)

        self.group_config_button = tk.Button(middle_frame, text='Group\nconfig', width=10, height=2, state=tk.DISABLED, command=self.group_config)
        self.group_config_button.grid(row=2, column=2, padx=20, pady=10)

        self.run_button = tk.Button(middle_frame, text='RUN', width=10, height=2, state=tk.DISABLED, command=self.run)
        self.run_button.grid(row=2, column=7, padx=20, pady=10)

        self.exp_button = tk.Button(middle_frame, text='Export', width=10, height=2, state=tk.DISABLED, command=self.export)
        self.exp_button.grid(row=2, column=8, padx=20, pady=10)

        self.open_exp_button = tk.Button(middle_frame, text='Open\nexport', width=10, height=2, state=tk.DISABLED, command=self.open_export)
        self.open_exp_button.grid(row=2, column=9, padx=20, pady=10)

    def get_std_path(self):
        path = filedialog.askopenfilename(title='Open standard file',
                                                   parent=self.root,
                                                   multiple=False,
                                                   filetype=[['std file', '.csv']])
        if path:
            self.std_path = path
            self.std_string.set(self.std_path)
            self.edit_std_button.config(state=tk.NORMAL)
            self._load()

    def get_data_path(self):
        path = filedialog.askopenfilename(title='Open data file',
                                                    parent=self.root,
                                                    multiple=False,
                                                    filetype=[['data file', '.xml']])
        if path:
            self.data_path = path
            self.data_string.set(self.data_path)
            self.edit_data_button.config(state=tk.NORMAL)
            self.new_std_button.config(state=tk.NORMAL)
            self.browse_std_button.config(state=tk.NORMAL)
            self.dir_button.config(state=tk.NORMAL)
            self.name_string.set(os.path.splitext(os.path.basename(path))[0])
            self._load()

    def _edit_std(self):
        if self.std_path != 'n/a':
            os.startfile(self.std_path)

    def _edit_data(self):
        if self.data_path != 'n/a':
            os.startfile(self.data_path)

    def _new_std(self):
        path = os.path.splitext(self.data_path)[0] + '_std' + '.csv'
        try:
            with open(path, 'w'):
                pass
        except PermissionError:
            messagebox.showerror('Error', 'Can\' create, file in use!')
            return

        os.startfile(path)
        self.edit_std_button.config(state=tk.NORMAL)
        self.std_path = path
        self.std_string.set(self.std_path)
        self._load()

    def _open_dir(self):
        path = os.path.split(self.data_path)[0].replace('/', '\\')
        subprocess.call('explorer {}'.format(path))

    def _load(self):
        self.group_config_button.config(state=tk.DISABLED)
        self.run_button.config(state=tk.DISABLED)
        self.exp_button.config(state=tk.DISABLED)
        self.open_exp_button.config(state=tk.DISABLED)

        if self.std_path != 'n/a' and self.data_path != 'n/a':
            # print('great!!!')
            self.load_button.config(state=tk.NORMAL)

    def load(self):
        if self.std_path != 'n/a' and self.data_path != 'n/a':
            print('great!!!')

            self.all = AllPlates()
            self.group_result = [0 for i in range(len(self.all.plates))]
            print(self.group_result)

            self.all.std_path = self.std_path
            self.all.data_path = self.data_path

            self.all.load_data()            # load data need to be above load std
            self.all.load_std()

            self.group_config_button.config(state=tk.NORMAL)
            self.run_button.config(state=tk.NORMAL)

    def group_config(self):

        dialog = GroupDialog(name_list=[i.name for i in self.all.plates],
                             result=self.group_result,
                             parent=self.root,
                             title='my dialog')

        if dialog.saved_data:
            self.group_result = dialog.saved_data
            print(dialog.group)
            self.all.set_group(dialog.group)

    def run(self):
        self.all.calc()
        self.all.rank()
        self.exp_button.config(state=tk.NORMAL)

    def export(self):
        while True:
            try:
                self.all.export()
                self.open_exp_button.config(state=tk.NORMAL)
                break
            except PermissionError:
                retry = messagebox.askretrycancel('Error', 'Export file in ues!!')
                if not retry:
                    break


    def open_export(self):
        os.startfile(self.all.excel_path)


class GroupDialog(simpledialog.Dialog):

    def __init__(self, name_list, result, *args, **kwargs):
        self.name_list = name_list
        self.color_dict = {0: '#ffffff', 1: '#eeeeee'}
        # self.groups = []
        self.col = len(result)
        self.length = len(name_list)
        self.group = []

        if result:
            print(result)
            self.saved_data = result
        else:
            # print('empty!')
            self.saved_data = []

        simpledialog.Dialog.__init__(self, *args, **kwargs)

        # print('init', self.saved_data)

    def body(self, master):
        print('result in body:\n', self.saved_data)
        main_frame = tk.Frame(master, width=800, height=500, padx=10, pady=10)
        main_frame.pack(side=tk.TOP, fill=tk.BOTH)

        top_frame = tk.Frame(main_frame, width=800, height=500)
        top_frame.pack(side=tk.TOP, anchor=tk.W)

        add = tk.Button(top_frame, text='Add a group', width=30, command=self.add_group)
        add.pack(side=tk.LEFT, anchor=tk.W)

        middle_frame = tk.Frame(main_frame)
        middle_frame.pack(side=tk.TOP, anchor=tk.W)

        self.list_frame = tk.Frame(middle_frame, width=300, height=500, bd=3, relief=tk.GROOVE, bg='#f6f7eb')
        self.list_frame.pack(side=tk.TOP, fill=tk.BOTH)

        for i in range(len(self.name_list)):
            tk.Label(self.list_frame, text=self.name_list[i], bg=self.color_dict[i%2]).grid(row=i, column=0, sticky=tk.E)
            tk.Label(self.list_frame, text=self.name_list[i], bg=self.color_dict[i%2]).grid(row=i, column=99, sticky=tk.W)

        # print('reslut!')
        # print(self.saved_data)

        if self.saved_data:
            # print('retrieving...')
            self.retrieve_group()
            # print('finished retrieving')

        limit = 6

        if len(self.name_list) < limit:
            limit = len(self.name_list)

        if len(self.saved_data) < limit:
            for _ in range(min(6 - len(self.saved_data), len(self.name_list))):
                self.add_group()

    def buttonbox(self):
        '''add standard button box.

        override if you do not want the standard buttons
        '''

        box = tk.Frame(self)

        w = tk.Button(box, text="Save and close", width=16, command=self.ok)
        w.pack(side=tk.LEFT, padx=5, pady=5)
        w = tk.Button(box, text="Clear and cancel", width=16, command=self.exit)
        w.pack(side=tk.LEFT, padx=5, pady=5)

        # self.bind("<Return>", self.ok)
        # self.bind("<Escape>", self.cancel)

        box.pack()

    def ok(self, event=None):

        if not self.validate():
            self.initial_focus.focus_set()  # put focus back
            return

        self.withdraw()
        self.update_idletasks()

        try:
            self.apply()
        finally:
            self.cancel()

    def cancel(self, event=None):

        # put focus back to the parent window
        if self.parent is not None:
            self.parent.focus_set()
        self.destroy()

    def exit(self):
        self.saved_data = []
        self.group = []
        self.cancel()

    def retrieve_group(self):

        for i in range(len(self.saved_data)):
            for j in range(len(self.saved_data[i])):
                tk.Checkbutton(self.list_frame, variable=self.saved_data[i][j], bg=self.color_dict[j%2]).grid(row=j, column=i+1)

    def add_group(self):

        if self.col > len(self.name_list):
            messagebox.showwarning('Warning!', 'Too many groups!!!', parent=self)
            return

        self.col += 1
        val_list = []
        for i in range(len(self.name_list)):
            val = tk.IntVar()
            tk.Checkbutton(self.list_frame, variable=val, bg=self.color_dict[i%2]).grid(row=i, column=self.col)
            val_list.append(val)
        self.saved_data.append(val_list)

        print(self.saved_data)

    def validate(self):
        # print('validation')

        # for i in self.saved_data:
        #     print([j.get() for j in i])

        if self.saved_data:
            err = []

            # print(len(self.saved_data))
            # print(len(self.saved_data[0]))
            # return 1

            self.matrix = list(map(list, zip(*self.saved_data)))

            for i in self.saved_data:
                print([j.get() for j in i])

            for i in range(len(self.matrix)):
                sss = sum([j.get() for j in self.matrix[i]])
                if sss > 1:
                    err.append(self.name_list[i])

            if err:
                errmsg = 'One plate in multiple groups:\n\n'
                for i in err:
                    errmsg = errmsg + i + '\n'
                messagebox.showerror('Error!!', errmsg)
                return 0
        return 1

    def apply(self):
        # print('apply')
        # print(self.matrix)

        self.group = []

        for i in range(len(self.name_list)):
            self.group.append(0)

        for i in range(len(self.matrix)):
            for j in range(len(self.matrix[i])):
                if self.matrix[i][j].get() == 1:
                    self.group[i] = j+1

        group_temp = self.group[:]
        print(group_temp)

        for i in range(len(self.group)):
            self.group[i] = 0

        counter = 1
        skip = []

        for i in range(len(group_temp)):
            if i in skip:
                continue

            if group_temp[i]:
                self.group[i] = counter
                skip.append(i)
            else:
                self.group[i] = 0
                continue

            try:
                for j in range(i+1, len(group_temp)):
                    if group_temp[j] == group_temp[i]:
                        self.group[j] = counter
                        skip.append(j)
            except IndexError:
                pass

            counter += 1

        print('self.group: ', len(self.group))
        print(self.group)

        print('names: ', len(self.name_list))
        print(self.name_list)




















































def test0():
    a = AllPlates()
    a.std_path = "D:\\Lifu\\OneDrive\\Shared\\Lab\\zy data\\412\\STD.csv"
    a.data_path = "D:\\Lifu\\OneDrive\\Shared\\Lab\\zy data\\412\\384.xml"

    a.load_std()
    a.load_data()



def test1():
    a = GUI()




















if __name__ == '__main__':
    test1()


